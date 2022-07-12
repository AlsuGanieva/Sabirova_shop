from openpyxl import Workbook
from openpyxl import load_workbook
from argparse import ArgumentParser
from argparse import FileType
from datetime import date
from candy_name_text_processor import Model
from hungarian_algorithm import algorithm
import os
import text_utils


class Candy:

    def __init__(self, art, name, unit, count=None, cost=None, summary=None):
        self.art = art
        self.name = name
        self.count = count
        self.unit = unit
        self.cost = cost
        self.summary = summary


def get_date():
    today = date.today()
    return today.strftime("%d-%m-%Y")


def generate_filename(output, name):
    return "{output}/{name}-{date}.xlsx".format(output=output, name=name, date=get_date())


def dir_path(string):
    if os.path.isdir(string):
        return string
    else:
        raise NotADirectoryError(string)


def init_args():
    parser = ArgumentParser(description="Накладные на кондитерку")
    parser.add_argument("-i1", "--input-1c", required=True, help="Выгрузка из 1С", type=FileType('r'))
    parser.add_argument("-iс", "--input-candy", required=True, help="Накладная \"кондитерка\"", type=FileType('r'))
    parser.add_argument("-o",
                        "--output",
                        required=False,
                        help="Папка, в которую сохранятся накладные",
                        type=dir_path,
                        default=".")
    return parser.parse_args()


def load_input_file(input_path):
    input_workbook = load_workbook(input_path)
    input_worksheet = input_workbook.active
    return input_worksheet


def read_1c_worksheet(input_worksheet):
    rows = []
    for row_number in range(2, input_worksheet.max_row):
        name = input_worksheet.cell(row_number, 6).value
        if name:
            rows.append(Candy(art=input_worksheet.cell(row_number, 3).value,
                              name=name,
                              unit=input_worksheet.cell(row_number, 4).value))
    return rows


def read_candy_worksheet(input_worksheet):
    rows = []
    for row_number in range(2, input_worksheet.max_row):
        name = input_worksheet.cell(row_number, 4).value
        if name:
            rows.append(Candy(art=input_worksheet.cell(row_number, 2).value,
                              name=name,
                              count=input_worksheet.cell(row_number, 5).value,
                              unit=input_worksheet.cell(row_number, 6).value,
                              cost=input_worksheet.cell(row_number, 7).value,
                              summary=input_worksheet.cell(row_number, 8).value))
    return rows


def generate_new_sheet(rows, title, file_name):
    output_workbook = Workbook()
    output_worksheet = output_workbook.active

    output_worksheet.title = title
    output_worksheet["A1"] = "арт"
    output_worksheet["B1"] = "штрих"
    output_worksheet["C1"] = "наименование"
    output_worksheet["D1"] = "Кол-во"
    output_worksheet["E1"] = "Ед.изм."
    output_worksheet["F1"] = "цена"
    output_worksheet["G1"] = "сумма"

    for row_index, row in enumerate(rows):
        output_worksheet.cell(row_index + 2, 1, value=row[0])
        output_worksheet.cell(row_index + 2, 3, value=row[1])
        output_worksheet.cell(row_index + 2, 4, value=row[2])
        output_worksheet.cell(row_index + 2, 5, value=row[3])
        output_worksheet.cell(row_index + 2, 6, value=row[4])
        output_worksheet.cell(row_index + 2, 7, value=row[5])
        if len(row) > 6:
            output_worksheet.cell(row_index + 2, 2, value=row[6])  # temp

    output_workbook.save(file_name)


def calculate_result(one_c_candies, candy_candies):
    identicals = 0
    likes = 0
    news = 0
    rows = []
    model = Model()
    model.fit(one_c_candies)
    for candy in candy_candies:
        if not text_utils.candy_name_split_regex.match(candy.name):
            raise ValueError("{} не по формату".format(candy.name))

        prediction, similarity, position, vector = model.predict(candy)
        similarity = round(similarity * 100)
        min_similarity = 85

        if prediction.art == candy.art and similarity >= min_similarity:
            # print("{} -> {} | {} IDENTICAL".format(prediction.name, candy.name, similarity))
            rows.append([candy.art, candy.name, candy.count, candy.unit, candy.cost, candy.summary,
                         "X{}".format(similarity)])
            identicals += 1
        elif similarity >= min_similarity:
            rows.append([candy.art, candy.name, candy.count, candy.unit, candy.cost, candy.summary])
            # print("{} -> {} | {} LIKE".format(prediction.name, candy.name, similarity).replace("\n", ""))
            likes += 1
        elif prediction.art == candy.art:
            rows.append([candy.art, candy.name, candy.count, candy.unit, candy.cost, candy.summary])
            print("{}|{} -> {}|{} | {} IDENTICAL".format(prediction.name, prediction.unit, candy.name, candy.unit, similarity))
            array_1c = model.result.toarray()
            array_v = vector.toarray()[0]
            for feature_name, point_1c, point in zip(model.pipeline.get_feature_names_out(), array_1c[position], array_v):
                if (point_1c > 0 or point > 0) and (point_1c != point):
                    print(feature_name, point_1c, point)
            likes += 1
        else:
            rows.append([candy.art, candy.name, candy.count, candy.unit, candy.cost, candy.summary])
            # print("NOT EXIST -> {} NOT EXIST".format(candy.name))
            news += 1
    print("IDENTICALS:{} | LIKES:{} | NEW ONES:{}".format(identicals, likes, news))
    return rows


if __name__ == '__main__':
    args = init_args()

    input_1c_worksheet = load_input_file(args.input_1c.name)
    input_candy_worksheet = load_input_file(args.input_candy.name)

    one_c_rows = read_1c_worksheet(input_1c_worksheet)
    candy_rows = read_candy_worksheet(input_candy_worksheet)

    rows = calculate_result(one_c_rows, candy_rows)
    generate_new_sheet(rows, "Кондитерка", generate_filename(args.output, "Кондитерка"))
