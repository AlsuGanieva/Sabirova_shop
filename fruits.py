from openpyxl import Workbook
from openpyxl import load_workbook
from argparse import ArgumentParser
from argparse import FileType
from datetime import date
import os
import re


class Shop:
    def __init__(self, name, column_number, percent, is_unformatted):
        self.name = name
        self.column_number = column_number
        self.percent = percent
        self.is_unformatted = is_unformatted


def load_input_file(input_path):
    input_workbook = load_workbook(input_path)
    input_worksheet = input_workbook.active
    name = input_worksheet["B2"].value
    return input_worksheet, name


def generate_new_sheet(rows, title, file_name):
    output_workbook = Workbook()
    output_worksheet = output_workbook.active

    output_worksheet.title = title
    output_worksheet["A1"] = "арт"
    output_worksheet["B1"] = "штрих"
    output_worksheet["C1"] = "наименование"
    output_worksheet["D1"] = "тара"
    output_worksheet["E1"] = "цена"
    output_worksheet["F1"] = "сумма"

    for row_index, row in enumerate(rows):
        output_worksheet.cell(row_index + 2, 1, value=row[0])
        output_worksheet.cell(row_index + 2, 3, value=row[1])
        output_worksheet.cell(row_index + 2, 4, value=strip_value(row[2]))
        output_worksheet.cell(row_index + 2, 5, value=round(strip_value(row[3]) / row[4]))
        output_worksheet.cell(row_index + 2, 6, value=strip_value(row[2]) * strip_value(row[3]))

    output_workbook.save(file_name)


def generate_unformatted_sheet(rows, title, file_name, percent):
    output_workbook = Workbook()
    output_worksheet = output_workbook.active

    output_worksheet.title = title
    output_worksheet["A1"] = "ТАРА"
    output_worksheet["B1"] = "НАИМЕНОВАНИЕ"
    output_worksheet["C1"] = "КОЛ-ВО"
    output_worksheet["D1"] = "ЦЕНА"
    output_worksheet["E1"] = "сумма"

    for row_index, row in enumerate(rows):
        output_worksheet.cell(row_index + 2, 1, value=row[2])
        output_worksheet.cell(row_index + 2, 2, value=row[1])
        output_worksheet.cell(row_index + 2, 4, value=round(row[3] / row[4] / 100.0 * percent))

    output_workbook.save(file_name)


def read_data(input_worksheet, column_number):
    rows = []
    for read_row in input_worksheet.iter_rows(min_row=3):
        count = read_row[column_number].value
        cost = read_row[5].value
        divide_by = read_row[6].value
        if divide_by:
            divide_by = float(divide_by)
        else:
            divide_by = 1.0
        if count and cost:
            rows.append([read_row[0].value,
                         read_row[1].value,
                         count,
                         cost,
                         divide_by])
    return rows


def get_date():
    today = date.today()
    return today.strftime("%d-%m-%Y")


def dir_path(string):
    if os.path.isdir(string):
        return string
    else:
        raise NotADirectoryError(string)


def init_args():
    parser = ArgumentParser(description="Накладные на фрукты и сухофрукты по магазинам")
    parser.add_argument("-i", "--input", required=True, help="Входная накладная с КУН И ценами", type=FileType('r'))
    parser.add_argument("-o",
                        "--output",
                        required=False,
                        help="Папка, в которую сохранятся накладные",
                        type=dir_path,
                        default=".")
    return parser.parse_args()


def generate_filename(output, name, title):
    return "{output}/{name}-{title}-{date}.xlsx".format(output=output, name=name, title=title, date=get_date())


def get_shops():
    return [
        Shop("Калинка", 2, 100, False),
        Shop("Удача", 3, 100, False),
        Shop("Надежда", 4, 127, True)
    ]


def strip_value(string):
    if not isinstance(string, str):
        return string
    result = re.search(r'^\D*(\d+[.,]?\d*)\D*$', string)
    if result:
        count = result.group(1)
        if count.isdigit():
            return int(count)
        else:
            return float(count)
    return None


if __name__ == '__main__':
    args = init_args()

    input_worksheet, name = load_input_file(args.input.name)
    for shop in get_shops():
        output_name = generate_filename(args.output, name, shop.name)
        rows = read_data(input_worksheet, shop.column_number)
        if shop.is_unformatted:
            generate_unformatted_sheet(rows, title=shop.name, file_name=output_name, percent=shop.percent)
        else:
            generate_new_sheet(rows, title=shop.name, file_name=output_name)
