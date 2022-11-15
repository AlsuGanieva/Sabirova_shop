import pyexcel  # pyexcel-xls pyexcel-xlsx
from openpyxl import Workbook
from openpyxl import load_workbook


class OneCRow:
    def __init__(self, art, code, name, count, cost, summary):
        self.art = art
        self.code = code
        self.name = name
        self.count = count
        self.cost = cost
        self.summary = summary


def generate_1c_sheet(oneCRows, title):
    output_workbook = Workbook()
    output_worksheet = output_workbook.active

    output_worksheet.title = title
    output_worksheet["A1"] = "арт"
    output_worksheet["B1"] = "штрих"
    output_worksheet["C1"] = "наименование"
    output_worksheet["D1"] = "тара"
    output_worksheet["E1"] = "цена"
    output_worksheet["F1"] = "сумма"

    for row_index, row in enumerate(oneCRows):
        output_worksheet.cell(row_index + 2, 1, value=row.art)
        output_worksheet.cell(row_index + 2, 2, value=row.code)
        output_worksheet.cell(row_index + 2, 3, value=row.name)
        output_worksheet.cell(row_index + 2, 4, value=row.count)
        output_worksheet.cell(row_index + 2, 5, value=row.cost)
        output_worksheet.cell(row_index + 2, 6, value=row.summary)

    apply_worksheet_width(output_worksheet)

    return output_workbook


def apply_worksheet_width(output_worksheet):
    dims = {}
    for row in output_worksheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        output_worksheet.column_dimensions[col].width = value + 3


def check_for_xls(path: str):
    if path.endswith(".xls"):
        new_path = path + "x"
        pyexcel.save_book_as(file_name=path, dest_file_name=new_path)
        return new_path
    else:
        return path


def load_input_workbook(input_path: str, handle_xls: bool = True) -> Workbook:
    if handle_xls:
        input_path = check_for_xls(input_path)
    input_workbook = load_workbook(input_path)
    return input_workbook
