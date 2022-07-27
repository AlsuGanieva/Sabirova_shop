import os
import re
from argparse import ArgumentParser
from argparse import FileType
from datetime import date

from openpyxl import Workbook

from utils import workbook_utils


class Shop:
    def __init__(self, name, column_number, percent, is_unformatted, should_join):
        self.name = name
        self.column_number = column_number
        self.percent = percent
        self.is_unformatted = is_unformatted
        self.should_join = should_join


class Fruit:
    def __init__(self, art, name, count, cost, divide_by):
        self.art = art
        self.name = name
        self.count = count
        self.cost = cost
        self.divide_by = divide_by


# todo: Move to workbook_utils
def generate_unformatted_sheet(fruits, title, file_name, percent):
    output_workbook = Workbook()
    output_worksheet = output_workbook.active

    output_worksheet.title = title
    output_worksheet["A1"] = "ТАРА"
    output_worksheet["B1"] = "НАИМЕНОВАНИЕ"
    output_worksheet["C1"] = "КОЛ-ВО"
    output_worksheet["D1"] = "ЦЕНА"
    output_worksheet["E1"] = "сумма"

    for row_index, fruit in enumerate(fruits):
        output_worksheet.cell(row_index + 2, 1, value=fruit.count)
        output_worksheet.cell(row_index + 2, 2, value=fruit.name)
        output_worksheet.cell(row_index + 2, 4, value=round((fruit.cost / fruit.divide_by) / 100.0 * percent))

    workbook_utils.apply_worksheet_width(output_worksheet)

    output_workbook.save(file_name)


def read_data(input_worksheet, column_number):
    fruits = []
    for read_row in input_worksheet.iter_rows(min_row=3):
        count = read_row[column_number].value
        cost = read_row[5].value
        divide_by = read_row[6].value
        if divide_by:
            divide_by = float(divide_by)
        else:
            divide_by = 1.0
        if count and cost:
            fruits.append(
                Fruit(art=read_row[0].value,
                      name=read_row[1].value,
                      count=count,
                      cost=cost,
                      divide_by=divide_by)
            )
    return fruits


def get_date():
    today = date.today()
    return today.strftime("%d-%m-%Y")


def dir_path(string):
    if os.path.isdir(string):
        return string
    else:
        raise NotADirectoryError(string)


def init_args():
    parser = ArgumentParser(description="Накладные на фрукты и сухофрукты по магазинам")
    parser.add_argument("-i", "--input", required=True, help="Входная накладная с КУН И ценами", type=FileType('r'))
    parser.add_argument("-o",
                        "--output",
                        required=False,
                        help="Папка, в которую сохранятся накладные",
                        type=dir_path,
                        default=".")
    return parser.parse_args()


def generate_filename(output, name, title):
    return "{output}/{name}-{title}-{date}.xlsx".format(output=output, name=name, title=title, date=get_date())


def get_shops():
    return [
        Shop("Калинка", 2, 100, is_unformatted=False, should_join=False),
        Shop("Удача", 3, 100, is_unformatted=False, should_join=True),
        Shop("Надежда", 4, 127, is_unformatted=True, should_join=True)
    ]


def strip_value(string):
    if not isinstance(string, str):
        return string
    result = re.search(r'^\D*(\d+[.,]?\d*)\D*$', string)
    if result:
        count = result.group(1)
        if count.isdigit():
            return int(count)
        else:
            return float(count)
    return None


def map_fruit_to_1c(fruits):
    oneCRows = []
    for fruit in fruits:
        cost = round(strip_value(fruit.cost) / fruit.divide_by)
        count = strip_value(fruit.count)
        oneCRows.append(
            workbook_utils.OneCRow(
                art=fruit.art,
                code="",
                name=fruit.name,
                count=count,
                cost=cost,
                summary=count * cost
            )
        )
    return oneCRows


def save_fruits(fruits, shop, output_name):
    if shop.is_unformatted:
        generate_unformatted_sheet(fruits, title=shop.name, file_name=output_name, percent=shop.percent)
    else:
        workbook = workbook_utils.generate_1c_sheet(map_fruit_to_1c(fruits), title=shop.name)
        workbook.save(output_name)


def process_fruits_facade(input_file_names, output_directory):
    concat_name = ""
    worksheets_and_names = []
    for file_name in input_file_names:
        input_worksheet = workbook_utils.load_input_worksheet(file_name)
        name = input_worksheet["B2"].value
        worksheets_and_names.append((input_worksheet, name))
        concat_name += name + "-"
    concat_name = concat_name[:-1]
    for shop in get_shops():
        fruits_for_shop = []
        for worksheet_and_name in worksheets_and_names:
            input_fruits = read_data(worksheet_and_name[0], shop.column_number)
            fruits_for_shop += input_fruits
            if not shop.should_join:
                output_name = generate_filename(output_directory, worksheet_and_name[1], shop.name)
                save_fruits(input_fruits, shop, output_name)
        if shop.should_join:
            output_name = generate_filename(output_directory, concat_name, shop.name)
            save_fruits(fruits_for_shop, shop, output_name)


if __name__ == '__main__':
    args = init_args()

    process_fruits_facade([args.input.name], args.output)
