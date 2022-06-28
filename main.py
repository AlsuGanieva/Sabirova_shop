from openpyxl import load_workbook
from openpyxl import Workbook

book = load_workbook('sweets.xlsx')
sheet = book.active
result_wb = Workbook()
result_ws = result_wb.active
for row in range(1, sheet.max_row):
    result_ws.append([sheet['B' + str(row)].value, sheet['C' + str(row)].value, sheet['D' + str(row)].value,
                      sheet['E' + str(row)].value, sheet['G' + str(row)].value, sheet['H' + str(row)].value])

result_wb.save("my_book.xlsx")
book.close()
