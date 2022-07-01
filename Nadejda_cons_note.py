from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Font, Border, Side)
from openpyxl.worksheet.page import PageMargins
import re

book = load_workbook('sweets.xlsx')
sheet_1 = book.active  # берем 1 лист активный
result_wb = Workbook()  # создаем рабочую книгу result_wb
result_ws = result_wb.worksheets[0]  # берем 1 лист
result_ws.title = 'Кондитерка'  # название 1 листа
name_cons_note = 'НАКЛАДНАЯ НА КОНДИТЕРКУ "НАДЕЖДА"'  # название накладной


def column_names():
    result_ws['A2'] = '№'
    result_ws['B2'] = 'Наименование'
    result_ws['C2'] = 'Кол-во'
    result_ws['D2'] = 'Ед. изм'
    result_ws['E2'] = 'Цена'
    result_ws['F2'] = 'Сумма'


column_names()

for row in range(2, sheet_1.max_row):
    result_ws.append(
        [row - 1, sheet_1[row][3].value, sheet_1[row][4].value, sheet_1[row][5].value,
         sheet_1[row][6].value, sheet_1[row][7].value])

thins = Side(border_style="thick", color="FF000000")  # жирный шрифт
double = Side(border_style="thin", color="FF000000")  # средний шрифт


def markup_calculation():  # функция расчета наценки на кондитерку
    for row in range(3, result_ws.max_row + 1):
        result_ws[row][4].value = round(1.27 * result_ws[row][4].value)  # процент


def text_alignment():  # функция выравнивания текста
    result_ws.page_setup.orientation = 'portrait'  # зададим размер и ориентацию страницы
    result_ws.page_setup.paperSize = result_ws.PAPERSIZE_A4
    cm = 1 / 2.54  # зададим собственные значения отступов
    result_ws.page_margins = PageMargins(left=1.5 * cm, right=cm, top=cm, bottom=cm)

    result_ws.merge_cells('A1:F1')
    result_ws['A1'] = name_cons_note  # наименование накладной
    result_ws['A1'].alignment = Alignment(horizontal="center", vertical="center")  # выравниваю название накл по центру
    result_ws['A1'].font = Font(bold=True, size=16)  # жирный шрифт, размер шрифта 16

    result_ws.column_dimensions['B'].width = 50  # изменяю ширину 2 колонки
    result_ws.column_dimensions['A'].width = 3  # изменяю ширину 1 колонки

    for row in range(3, result_ws.max_row + 1):  # перенос текста, если не помещается
        result_ws[row][1].alignment = Alignment(horizontal='general', vertical="center")
        result_ws[row][4].font = Font(bold=True, size=14)  # жирный шрифт, размер шрифта 14
        if len(str(result_ws[row][1].value)) > 50:
            result_ws.row_dimensions[row].height = 40
            result_ws[row][1].alignment = Alignment(wrapText=True, horizontal='general', vertical="center")

    for cell in range(6):  # центрирование названия колонок
        result_ws[2][cell].alignment = Alignment(horizontal="center", vertical="center")

    for cell in range(2, 6):  # центрирование колонок
        for row in range(3, result_ws.max_row + 1):
            result_ws[row][cell].alignment = Alignment(horizontal="center", vertical="center")
            result_ws[row][0].alignment = Alignment(horizontal="center", vertical="center")

    for cell in range(6):  # делаю сетку
        for row in range(2, result_ws.max_row + 1):
            result_ws[row][cell].border = Border(top=double, bottom=double, left=double, right=double)


def summary_output():  # функция подсчета итоговых сумм
    m = int(result_ws.max_row)  # максимальная ячейка
    # выведем сумму накладной, выделим жирным шрифтом и обведем ячейку
    summa = 0
    for row in range(3, m + 1):
        summa += result_ws[row][5].value
    result_ws[m + 1][5].value = round(summa)
    result_ws[m + 1][5].alignment = Alignment(horizontal="center", vertical="center")
    result_ws[m + 1][5].border = Border(top=double, bottom=double, left=double, right=double)

    # выведем итог для отчета
    total = round(summa * 1.27)
    result_ws[m + 2][1].value = 'ИТОГ ДЛЯ ОТЧЕТА:  {} руб'.format(total)
    result_ws[m + 2][1].alignment = Alignment(horizontal="center", vertical="center")
    result_ws[m + 2][1].border = Border(top=thins, bottom=thins, left=thins, right=thins)
    result_ws[m + 2][1].font = Font(bold=True, size=14)  # жирный шрифт, размер шрифта 14


markup_calculation()  # вызов функции расчета наценки на кондитерку
text_alignment()  # вызов функции выравнивания текста
summary_output()  # вызов функции подсчета итогов

# работа с колбасой 7 склад

book_2 = load_workbook('sausage7.xlsx')
sheet_2 = book_2.worksheets[2]  # берем 1 лист активный
result_wb.create_sheet(title='Колбаса', index=1)  # Добавить новый лист в Excel
result_ws = result_wb.worksheets[1]  # берем 1 лист
name_cons_note = 'НАКЛАДНАЯ НА КОЛБАСУ "НАДЕЖДА"'  # название накладной

column_names()

for row in range(3, sheet_2.max_row + 1):
    result_ws.append([row - 2, sheet_2[row][0].value, sheet_2[row][1].value, '',
                      sheet_2[row][2].value, sheet_2[row][3].value])

book_3 = load_workbook('sausage75.xlsx')
sheet_3 = book_3.worksheets[0]  # берем 1 лист активный


def strip_value(string):
    if len(string) > 0:
        result = re.sub(r"['\s]", '', string)
        return float(result)
    else:
        return None


for row in range(result_ws.max_row + 1, sheet_3.max_row - 9):
    result_ws.append([result_ws.max_row + 1, sheet_3[row][1].value, strip_value(sheet_3[row][3].value),
                      sheet_3[row][5].value, strip_value(sheet_3[row][2].value),
                      strip_value(sheet_3[row][6].value)])

markup_calculation()  # вызов функции расчета наценки
text_alignment()  # вызов функции выравнивания текста
summary_output()  # вызов функции подсчета итогов

# Герасимова

book_4 = load_workbook('Gerasimova.xlsx')
sheet_4 = book_4.worksheets[0]  # берем 1 лист активный
result_wb.create_sheet(title='Герасимова', index=2)  # Добавить новый лист в Excel
result_ws = result_wb.worksheets[2]  # берем 3 лист
name_cons_note = 'НАКЛАДНАЯ ГЕРАСИМОВА "НАДЕЖДА"'  # название накладной

column_names()

for row in range(13, sheet_4.max_row - 7):
    result_ws.append([row - 2, sheet_4[row][3].value, sheet_4[row][20].value,
                      sheet_4[row][23].value, sheet_4[row][25].value,
                      sheet_4[row][29].value])


def is_has_substring(row, substring):  # функция поиска подстроки
    result = re.findall(substring, result_ws[row][1].value)
    return len(result) != 0


def markup_calculation_gerasimova():  # функция расчета наценки Герасимова
    for row in range(3, result_ws.max_row + 1):
        if is_has_substring(row, 'ПЕСОК'):
            result_ws[row][4].value = round(1.2 * result_ws[row][4].value)
        elif is_has_substring(row, 'КИТЕКАТ СУХ. 15КГ'):
            result_ws[row][2].value *= 15
            result_ws[row][3].value = 'кг'
            result_ws[row][4].value = round(1.27 * (result_ws[row][4].value / 15))
        elif result_ws[row][4].value <= 8:
            result_ws[row][4].value += 2
            result_ws[row][4].value = round(result_ws[row][4].value)
        else:
            result_ws[row][4].value = round(1.27 * result_ws[row][4].value)


markup_calculation_gerasimova()  # вызов функции расчета наценки
text_alignment()  # вызов функции выравнивания текста
summary_output()  # вызов функции подсчета итогов

result_wb.save("my_book.xlsx")
book.close()
